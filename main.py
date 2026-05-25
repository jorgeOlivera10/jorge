from fastapi import FastAPI, HTTPException, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from .models import ParseRequest, ParseResponse, CalculateRequest, CalculateResponse, CalculateItemResponse, ExportRequest
from .calculator import (
    parse_input, 
    calculate_cable, 
    calculate_inter_hall_cable, 
    validate_pair_for_calc,
    apply_rounding
)

from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Calculadora AWS Backend")

# Allow CORS for the Vite dev server (http://localhost:5173) and any other origin during development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/api/parse", response_model=ParseResponse)
def api_parse(req: ParseRequest):
    pairs = parse_input(req.text)
    return ParseResponse(pairs=pairs)

@app.post("/api/calculate", response_model=CalculateResponse)
def api_calculate(req: CalculateRequest):
    results = []
    has_errors = False

    for item in req.items:
        pair = item.pair.model_dump()
        tier = item.tier
        path = item.path or "A"
        rounding_mode = item.roundingMode
        inter_hall_meters = item.interHallMeters

        if pair.get("error"):
            results.append(CalculateItemResponse(
                pair=item.pair,
                roundingMode=rounding_mode,
                pathLabel=path,
                error=pair["error"]
            ))
            has_errors = True
            continue

        if pair.get("interHall"):
            if inter_hall_meters is None or inter_hall_meters < 0:
                results.append(CalculateItemResponse(
                    pair=item.pair,
                    roundingMode=rounding_mode,
                    pathLabel=path,
                    error=f'Error en {pair["origin"]["raw"]} -> {pair["dest"]["raw"]}: Introduce la distancia Inter-Hall.'
                ))
                has_errors = True
                continue
            
            calc = calculate_inter_hall_cable(pair["origin"], pair["dest"], path, tier, inter_hall_meters)
            if calc.get("error"):
                results.append(CalculateItemResponse(
                    pair=item.pair,
                    roundingMode=rounding_mode,
                    pathLabel=path,
                    error=f'Origen: {pair["origin"]["raw"]} / Destino: {pair["dest"]["raw"]} - {calc["error"]}'
                ))
                has_errors = True
                continue
            
            calc["totalRaw"] = calc["total"]
            calc["total"] = apply_rounding(calc["total"], rounding_mode)
            results.append(CalculateItemResponse(
                pair=item.pair,
                roundingMode=rounding_mode,
                pathLabel=path,
                calc=calc
            ))
            continue

        if pair.get("directConnection"):
            calc = calculate_cable(pair["origin"], pair["dest"], None, tier)
            calc["interHallDist"] = 0.0
            calc["totalRaw"] = calc["total"]
            calc["total"] = apply_rounding(calc["total"], rounding_mode)
            results.append(CalculateItemResponse(
                pair=item.pair,
                roundingMode=rounding_mode,
                pathLabel="Directa",
                calc=calc
            ))
            continue

        validation_errors = validate_pair_for_calc(pair, path, tier)
        if validation_errors:
            results.append(CalculateItemResponse(
                pair=item.pair,
                roundingMode=rounding_mode,
                pathLabel=path,
                error=f'Origen: {pair["origin"]["raw"]} / Destino: {pair["dest"]["raw"]} - {"; ".join(validation_errors)}'
            ))
            has_errors = True
            continue
        
        calc = calculate_cable(pair["origin"], pair["dest"], path, tier)
        if calc.get("error"):
            results.append(CalculateItemResponse(
                pair=item.pair,
                roundingMode=rounding_mode,
                pathLabel=path,
                error=f'Origen: {pair["origin"]["raw"]} / Destino: {pair["dest"]["raw"]} - {calc["error"]}'
            ))
            has_errors = True
            continue

        calc["totalRaw"] = calc["total"]
        calc["total"] = apply_rounding(calc["total"], rounding_mode)
        results.append(CalculateItemResponse(
            pair=item.pair,
            roundingMode=rounding_mode,
            pathLabel=path,
            calc=calc
        ))

    return CalculateResponse(results=results, hasErrors=has_errors)

@app.post("/api/export")
def api_export(req: ExportRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cableado"

    # Determinar si hay redondeo para ajustar las columnas
    has_any_rounding = any(r.roundingMode and r.roundingMode != "0" for r in req.results)

    # Encabezados
    if has_any_rounding:
        headers = ['Rack Origen', 'Rack Destino', 'Path', 'Vertical Inicial (m)', 'Horizontal (m)', 'Inter-Hall (m)', 'Vertical Final (m)', 'Bandeja (m)', 'Margen (m)', 'Total Exacto (m)', 'Total Redondeado (m)']
    else:
        headers = ['Rack Origen', 'Rack Destino', 'Path', 'Vertical Inicial (m)', 'Horizontal (m)', 'Inter-Hall (m)', 'Vertical Final (m)', 'Bandeja (m)', 'Margen (m)', 'Total (m)']

    ws.append(headers)

    # Estilos de encabezado
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin")
    header_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # Datos
    for r in req.results:
        p = r.pair
        c = r.calc
        if not c:
            continue

        ih_val = c.interHallDist if c.interHallDist > 0 else 0
        
        row = [
            p.origin.raw if p.origin else "Error",
            p.dest.raw if p.dest else "Error",
            r.pathLabel,
            c.verticalInicial,
            c.horizontal,
            ih_val,
            c.verticalFinal,
            c.bandeja,
            c.margen
        ]

        if has_any_rounding:
            raw_val = c.totalRaw if c.totalRaw is not None else c.total
            row.extend([raw_val, c.total])
        else:
            row.append(c.total)
        
        ws.append(row)

    # Totales
    total_meters = sum(r.calc.total for r in req.results if r.calc)
    total_raw_sum = sum((r.calc.totalRaw if r.calc.totalRaw is not None else r.calc.total) for r in req.results if r.calc)
    
    total_row_idx = len(req.results) + 2
    if has_any_rounding:
        total_row = ['TOTAL', '', '', '', '', '', '', '', '', total_raw_sum, total_meters]
    else:
        total_row = ['TOTAL', '', '', '', '', '', '', '', '', total_meters]
    
    ws.append(total_row)
    
    # Estilos para celdas de datos y totales
    for row_idx in range(2, ws.max_row + 1):
        is_total_row = (row_idx == ws.max_row)
        for cell_idx, cell in enumerate(ws[row_idx], 1):
            cell.border = header_border
            if is_total_row:
                cell.font = header_font
            
            # Formato numérico para columnas de metros (desde la 4 en adelante)
            if cell_idx >= 4 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Ajustar anchos de columna (aproximado)
    column_widths = [20, 20, 10, 18, 15, 14, 18, 13, 13, 16, 18] if has_any_rounding else [20, 20, 10, 18, 15, 14, 18, 13, 13, 13]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    excel_data = output.getvalue()
    output.close()

    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Cableado_ZAZ.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

frontend_dir = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(frontend_dir):
    app.mount("/static", StaticFiles(directory=frontend_dir), name="static")

    @app.get("/")
    def read_index():
        return FileResponse(os.path.join(frontend_dir, "index.html"))

    @app.get("/{path:path}")
    def catch_all(path: str):
        file_path = os.path.join(frontend_dir, path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(frontend_dir, "index.html"))
